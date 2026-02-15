from __future__ import annotations

import re
import traceback
from io import BytesIO
from datetime import datetime
from typing import Dict, List, Tuple, Optional

import pandas as pd
from openpyxl import load_workbook
from google.cloud import storage
from google.cloud import bigquery

# =========================
# CONFIG
# =========================
PROJECT_ID = "gcp-sfpl-etl-demo-project6"
DATASET_ID = "Shravas_Lab"
BUCKET_NAME = "shravas-lab"
GCS_PREFIX = "DATA/"

HEADER_ROW_1BASED = 6
SKIP_FIRST_SHEET = True
PANDAS_ENGINE = "openpyxl"

EXCLUDE_MAINMETRICS = {
    "RatiosMetrics",
     "GeneralMetrics",
     "StockMetrics",
     "FinanceMetrics",
     #"EquityMetrics",
}
EXCLUDE_SUBMETRICS ={
'BonusHistory',
'CapitalHistory',
'DividendHistory',
'FCCB',
#'HoldingGreaterThan1%Promoters',
#'HoldingGreaterThan1%Public',
'MFHoldings',
#'PromotersHolding',
'ShareHoldingPattern',
}

DROP_COLUMNS_REGEX = [r"^unnamed"]

CHUNK_SIZE_ROWS = 500000

# Output period format:
# False -> YYMM like 2509
# True  -> YYYYMM like 202509
FORCE_YYYYMM = False

# =========================
# Helpers
# =========================
MONTH_MAP = {
    "jan": 1, "january": 1,
    "feb": 2, "february": 2,
    "mar": 3, "march": 3,
    "apr": 4, "april": 4,
    "may": 5,
    "jun": 6, "june": 6,
    "jul": 7, "july": 7,
    "aug": 8, "august": 8,
    "sep": 9, "sept": 9, "september": 9,
    "oct": 10, "october": 10,
    "nov": 11, "november": 11,
    "dec": 12, "december": 12,
}

def log(msg: str) -> None:
    print(msg, flush=True)

def sanitize_col(name: str) -> str:
    name = "" if name is None else str(name)
    name = name.strip().replace("\n", " ").replace("\r", " ")
    name = re.sub(r"[^\w]+", "_", name)
    name = re.sub(r"_+", "_", name).strip("_")
    if not name:
        name = "col"
    if name[0].isdigit():
        name = f"c_{name}"
    return name.lower()

def make_unique(cols: List[str]) -> List[str]:
    seen: Dict[str, int] = {}
    out = []
    for c in cols:
        if c not in seen:
            seen[c] = 1
            out.append(c)
        else:
            seen[c] += 1
            out.append(f"{c}_{seen[c]}")
    return out

def safe_table_name(mainmetric: str, submetric: str) -> str:
    base = f"{mainmetric}_{submetric}"
    base = re.sub(r"[^\w]+", "_", base).strip("_")
    if not base:
        base = "table"
    if base[0].isdigit():
        base = f"t_{base}"
    return base

def derive_mainmetric_submetric(blob_name: str) -> Optional[Tuple[str, str]]:
    if not blob_name.lower().endswith(".xlsx"):
        return None
    parts = blob_name.split("/")
    if len(parts) < 3:
        return None
    prefix = GCS_PREFIX.strip("/")
    try:
        data_idx = parts.index(prefix)
    except ValueError:
        data_idx = 0
    if data_idx + 1 >= len(parts):
        return None
    mainmetric = parts[data_idx + 1].strip()
    filename = parts[-1].strip()
    submetric = filename[:-5]
    if not mainmetric or not submetric:
        return None
    return mainmetric, submetric

def workbook_sheet_max_rows(excel_bytes: bytes) -> Dict[str, int]:
    wb = load_workbook(BytesIO(excel_bytes), read_only=True, data_only=True)
    out = {name: int(wb[name].max_row or 0) for name in wb.sheetnames}
    wb.close()
    return out

def drop_junk_columns(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    out.dropna(axis=1, how="all", inplace=True)
    out.dropna(axis=0, how="all", inplace=True)
    regexes = [re.compile(pat, flags=re.IGNORECASE) for pat in DROP_COLUMNS_REGEX]
    cols_to_drop = [c for c in out.columns if any(rgx.search(c) for rgx in regexes)]
    if cols_to_drop:
        out.drop(columns=cols_to_drop, inplace=True)
    return out

def to_str_series(s: pd.Series) -> pd.Series:
    return s.map(lambda x: None if pd.isna(x) else str(x))

# =========================
# Period parsing
# =========================
def parse_month_header_to_yymm(s: str) -> Optional[str]:
    """
    Parses strings like Sep-25, Sep'25, Sep 2025, 2025-09, 09-2025
    Returns yymm as YYMM or YYYYMM based on FORCE_YYYYMM.
    """
    if s is None:
        return None
    s = str(s).strip()
    if not s:
        return None

    s_clean = s.lower().replace("_", " ").strip()
    s_clean = re.sub(r"\s+", " ", s_clean)

    # YYYY-MM
    m = re.match(r"^(\d{4})[-/ ](\d{1,2})$", s_clean)
    if m:
        y = int(m.group(1)); mo = int(m.group(2))
        if 1 <= mo <= 12:
            return f"{y:04d}{mo:02d}" if FORCE_YYYYMM else f"{y%100:02d}{mo:02d}"

    # MM-YYYY
    m = re.match(r"^(\d{1,2})[-/ ](\d{4})$", s_clean)
    if m:
        mo = int(m.group(1)); y = int(m.group(2))
        if 1 <= mo <= 12:
            return f"{y:04d}{mo:02d}" if FORCE_YYYYMM else f"{y%100:02d}{mo:02d}"

    # Mon-YY / Mon'YY / Mon YYYY
    m = re.match(r"^([a-z]{3,9})[ \-'](\d{2}|\d{4})$", s_clean)
    if m:
        mon_txt = m.group(1); yy = m.group(2)
        if mon_txt in MONTH_MAP:
            mo = MONTH_MAP[mon_txt]
            if len(yy) == 2:
                y2 = int(yy)
                y = 2000 + y2 if y2 <= 69 else 1900 + y2
            else:
                y = int(yy)
            return f"{y:04d}{mo:02d}" if FORCE_YYYYMM else f"{y%100:02d}{mo:02d}"

    # Mon YYYY
    m = re.match(r"^([a-z]{3,9}) (\d{4})$", s_clean)
    if m:
        mon_txt = m.group(1); y = int(m.group(2))
        if mon_txt in MONTH_MAP:
            mo = MONTH_MAP[mon_txt]
            return f"{y:04d}{mo:02d}" if FORCE_YYYYMM else f"{y%100:02d}{mo:02d}"

    return None

SUFFIX_PERIOD_RE = re.compile(r"^(?P<base>.+?)[ _-]*(?P<yymm>\d{4})$")

def parse_yymm4(yymm4: str) -> Optional[str]:
    """
    Validates YYMM like 2512 and returns YYMM or YYYYMM.
    """
    if not re.fullmatch(r"\d{4}", yymm4):
        return None
    yy = int(yymm4[:2])
    mm = int(yymm4[2:])
    if not (1 <= mm <= 12):
        return None
    if FORCE_YYYYMM:
        y = 2000 + yy if yy <= 69 else 1900 + yy
        return f"{y:04d}{mm:02d}"
    return f"{yy:02d}{mm:02d}"

# =========================
# NEW: handle "Sep-25\nNo of Shares" style headers
# =========================
def split_multiline_header(header: str) -> Tuple[Optional[str], str]:
    """
    Given a header cell string that may contain newlines, attempt to extract:
      - yymm (from any line that looks like a period)
      - base_name (from remaining text)
    Returns (yymm_or_None, base_name_sanitized)
    """
    raw = "" if header is None else str(header)
    parts = [p.strip() for p in re.split(r"[\n\r]+", raw) if p and p.strip()]

    # If no newline parts, just return as base (no yymm found here)
    if not parts:
        return None, sanitize_col(raw)

    # Find a line that parses as a month period
    yymm = None
    yymm_idx = None
    for i, p in enumerate(parts):
        cand = parse_month_header_to_yymm(p)
        if cand:
            yymm = cand
            yymm_idx = i
            break

    # Base name = join of remaining parts excluding yymm part
    if yymm is not None:
        base_parts = [p for j, p in enumerate(parts) if j != yymm_idx]
        base = " ".join(base_parts).strip()
        if not base:
            base = "value"
        return yymm, sanitize_col(base)

    # If nothing parses as month, return whole as base
    return None, sanitize_col(" ".join(parts))

# =========================
# Core transform:
# - remove time from column names
# - create single yymm column
# - keep base column names (wide measures)
# =========================
def normalize_period_columns_to_yymm(df: pd.DataFrame) -> pd.DataFrame:
    """
    Submetric-neutral. Handles:
    A) suffix style: <base>_2512
    B) pure month columns: Sep-25, Jan-24
    C) multiline headers: "Sep-25\\nNo of Shares"  (period + measure in same cell)
    Output: period removed from col names and a single yymm column added.
    """
    df = df.copy()

    # --- Classify each column into:
    # 1) suffix-period measure column
    # 2) pure-month column (no measure)
    # 3) multiline header column (period + measure)
    suffix_cols = []
    suffix_parsed: List[Tuple[str, str, str]] = []  # (orig_col, base, yymm)

    multiline_cols = []
    multiline_parsed: List[Tuple[str, str, str]] = []  # (orig_col, base, yymm)

    pure_month_map: Dict[str, str] = {}  # col -> yymm

    for c in df.columns:
        if c == "companyID":
            continue

        # (C) multiline header check FIRST
        if isinstance(c, str) and ("\n" in c or "\r" in c):
            yymm, base = split_multiline_header(c)
            if yymm:
                multiline_cols.append(c)
                multiline_parsed.append((c, base, yymm))
                continue  # don't treat further

        # (A) suffix period like <base>_2512
        m = SUFFIX_PERIOD_RE.match(str(c))
        if m:
            yymm = parse_yymm4(m.group("yymm"))
            if yymm:
                base = sanitize_col(m.group("base"))
                suffix_cols.append(c)
                suffix_parsed.append((c, base, yymm))
                continue

        # (B) pure month header column like Sep-25
        yymm = parse_month_header_to_yymm(str(c))
        if yymm:
            pure_month_map[c] = yymm
            continue

    if not suffix_parsed and not pure_month_map and not multiline_parsed:
        return df  # nothing to normalize

    out_frames: List[pd.DataFrame] = []

    # Helper to pivot parsed measure columns (suffix or multiline) into wide by yymm
    def pivot_measure_columns(parsed_cols: List[Tuple[str, str, str]], cols_to_remove: List[str]) -> pd.DataFrame:
        id_cols = [c for c in df.columns if c not in cols_to_remove]
        parts = []
        for orig_col, base, yymm in parsed_cols:
            tmp = df[id_cols].copy()
            tmp["yymm"] = yymm
            tmp["_base"] = base
            tmp["_value"] = df[orig_col]
            parts.append(tmp)

        long_df = pd.concat(parts, ignore_index=True)

        wide = (
            long_df.pivot_table(
                index=id_cols + ["yymm"],
                columns="_base",
                values="_value",
                aggfunc="first",
            )
            .reset_index()
        )
        wide.columns = make_unique([sanitize_col(x) for x in wide.columns])
        return wide

    # Combine suffix + multiline measure columns into one pivot (same structure)
    combined_measure_parsed = suffix_parsed + multiline_parsed
    combined_measure_cols = suffix_cols + multiline_cols

    if combined_measure_parsed:
        out_frames.append(pivot_measure_columns(combined_measure_parsed, combined_measure_cols))

    # Pure month columns => melt to yymm + value
    if pure_month_map:
        df2 = df.copy()

        inv: Dict[str, List[str]] = {}
        for col, yymm in pure_month_map.items():
            inv.setdefault(yymm, []).append(col)

        yymm_cols = []
        for yymm, cols in inv.items():
            if len(cols) == 1:
                df2.rename(columns={cols[0]: yymm}, inplace=True)
                yymm_cols.append(yymm)
            else:
                tmp = df2[cols].copy()
                df2[yymm] = tmp.bfill(axis=1).iloc[:, 0]
                df2.drop(columns=cols, inplace=True)
                yymm_cols.append(yymm)

        id_cols2 = [c for c in df2.columns if c not in yymm_cols]
        long = df2.melt(
            id_vars=id_cols2,
            value_vars=sorted(set(yymm_cols)),
            var_name="yymm",
            value_name="value",
        )
        long.dropna(subset=["value"], how="all", inplace=True)
        long.columns = make_unique([sanitize_col(x) for x in long.columns])
        out_frames.append(long)

    # Merge frames if more than one pattern exists
    result = out_frames[0]
    for nxt in out_frames[1:]:
        # merge keys: companyID + yymm + any common id cols
        keys = [c for c in ["companyid", "yymm"] if c in result.columns and c in nxt.columns]
        common = set(result.columns).intersection(set(nxt.columns))
        for c in sorted(common):
            if c in keys:
                continue
            if c in {"value"}:
                continue
            keys.append(c)
        result = result.merge(nxt, on=keys, how="outer")

    return result

# =========================
# Excel read
# =========================
def read_excel_sheets(excel_bytes: bytes, source_label: str) -> List[pd.DataFrame]:
    dfs: List[pd.DataFrame] = []
    try:
        max_rows = workbook_sheet_max_rows(excel_bytes)
    except Exception as e:
        log(f"⚠️ Cannot precheck workbook rows: source={source_label} err={e}")
        max_rows = {}

    with pd.ExcelFile(BytesIO(excel_bytes), engine=PANDAS_ENGINE) as xls:
        sheet_names = xls.sheet_names
        if SKIP_FIRST_SHEET and sheet_names:
            sheet_names = sheet_names[1:]

        for sh in sheet_names:
            if sh in max_rows and max_rows[sh] < HEADER_ROW_1BASED:
                continue
            try:
                df = pd.read_excel(
                    xls,
                    sheet_name=sh,
                    header=HEADER_ROW_1BASED - 1,
                    dtype=object
                )

                df.dropna(axis=0, how="all", inplace=True)
                df.dropna(axis=1, how="all", inplace=True)
                if df.empty:
                    continue

                df.insert(0, "companyID", str(sh).strip())

                # IMPORTANT: keep original headers (including newlines) until after classification
                # sanitize AFTER pandas reads them
                df.columns = make_unique([c if isinstance(c, str) else str(c) for c in df.columns])

                # drop junk columns (works on unsanitized too)
                df.columns = [str(c) for c in df.columns]
                df = drop_junk_columns(df)
                if df.empty:
                    continue

                # Now sanitize columns, but preserve newline info by converting later in normalize function
                # We keep as-is for normalize_period_columns_to_yymm, then sanitize output columns there.
                dfs.append(df)

            except Exception as e:
                log(f"⚠️ Sheet read failed: source={source_label} sheet={sh} err={e}")
                continue

    return dfs

# =========================
# BigQuery load (safe: no pyarrow type fights)
# =========================
def load_df_to_bq_safe(bq: bigquery.Client, df: pd.DataFrame, table_id: str, first_load: bool) -> None:
    """
    Cast everything to STRING before upload to avoid inconsistent-type pyarrow errors.
    (Does not add columns; only casts.)
    """
    if df.empty:
        return

    df2 = df.copy()
    # sanitize final column names (ensures BQ-safe)
    df2.columns = make_unique([sanitize_col(c) for c in df2.columns])

    for c in df2.columns:
        df2[c] = to_str_series(df2[c])

    write_disp = bigquery.WriteDisposition.WRITE_TRUNCATE if first_load else bigquery.WriteDisposition.WRITE_APPEND
    job_config = bigquery.LoadJobConfig(
        autodetect=True,
        write_disposition=write_disp,
        schema_update_options=[bigquery.SchemaUpdateOption.ALLOW_FIELD_ADDITION] if not first_load else None
    )

    job = bq.load_table_from_dataframe(df2, table_id, job_config=job_config)
    job.result()

# =========================
# Main
# =========================
def main():
    storage_client = storage.Client(project=PROJECT_ID)
    bq_client = bigquery.Client(project=PROJECT_ID)

    bucket = storage_client.bucket(BUCKET_NAME)
    blobs = list(bucket.list_blobs(prefix=GCS_PREFIX))
    log(f"Found {len(blobs)} blobs under gs://{BUCKET_NAME}/{GCS_PREFIX}")

    groups: Dict[Tuple[str, str], List[storage.Blob]] = {}
    for blob in blobs:
        derived = derive_mainmetric_submetric(blob.name)
        if not derived:
            continue
        mainmetric, submetric = derived
        if mainmetric in EXCLUDE_MAINMETRICS or submetric in EXCLUDE_SUBMETRICS:
            continue
        groups.setdefault((mainmetric, submetric), []).append(blob)

    log(f"Grouped into {len(groups)} tables after exclusions")

    for (mainmetric, submetric), group_blobs in sorted(groups.items()):
        table_id = f"{PROJECT_ID}.{DATASET_ID}.{safe_table_name(mainmetric, submetric)}"
        log(f"\n=== {mainmetric}.{submetric} -> {table_id} ===")

        group_blobs = sorted(group_blobs, key=lambda b: b.name)
        first_load = True

        for blob in group_blobs:
            try:
                excel_bytes = blob.download_as_bytes()
                sheet_dfs = read_excel_sheets(excel_bytes, source_label=blob.name)
                if not sheet_dfs:
                    continue

                file_df = pd.concat(sheet_dfs, ignore_index=True)
                if file_df.empty:
                    continue

                out_df = normalize_period_columns_to_yymm(file_df)
                if out_df.empty:
                    continue

                total = len(out_df)
                for start in range(0, total, CHUNK_SIZE_ROWS):
                    end = min(start + CHUNK_SIZE_ROWS, total)
                    chunk = out_df.iloc[start:end].copy()

                    write_first = first_load and (start == 0)
                    load_df_to_bq_safe(bq_client, chunk, table_id, first_load=write_first)

                    mode = "WRITE_TRUNCATE" if write_first else "WRITE_APPEND"
                    log(f"  -> LOADED {blob.name.split('/')[-1]} rows={len(chunk)} mode={mode}")

                first_load = False

            except KeyboardInterrupt:
                log("\n⛔ Stopped by user (KeyboardInterrupt). Exiting cleanly.")
                return
            except Exception as e:
                log(f"❌ ERROR blob={blob.name}: {e}")
                traceback.print_exc()

    log("\n✅ Done.")

if __name__ == "__main__":
    main()
